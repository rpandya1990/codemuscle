import csv
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from codemuscle.domain.exceptions import ImportFileError


def read_tabular_file(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    if path.suffix.casefold() == ".csv":
        return _read_csv(path)
    if path.suffix.casefold() == ".xlsx":
        return _read_xlsx(path)
    raise ImportFileError("Only .csv and .xlsx files are supported.")


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = [header.strip() for header in (reader.fieldnames or []) if header]
            rows: list[dict[str, object]] = []
            for row in reader:
                parsed_row: dict[str, object] = {}
                for key, value in row.items():
                    if key is not None:
                        parsed_row[str(key).strip()] = value
                rows.append(parsed_row)
    except (UnicodeDecodeError, csv.Error) as error:
        raise ImportFileError("The CSV file could not be parsed.") from error
    if not headers:
        raise ImportFileError("The import file has no header row.")
    return headers, rows


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if not workbook.worksheets:
            raise ImportFileError("The workbook is empty.")
        worksheet = workbook.worksheets[0]
        values = worksheet.iter_rows(values_only=True)
        first_row = next(values, None)
        if first_row is None:
            raise ImportFileError("The workbook is empty.")
        headers = [str(value).strip() if value is not None else "" for value in first_row]
        if not any(headers):
            raise ImportFileError("The workbook has no header row.")
        rows = [
            {
                header: _json_value(value)
                for header, value in zip(headers, row, strict=False)
                if header
            }
            for row in values
        ]
        workbook.close()
        return headers, rows
    except ImportFileError:
        raise
    except Exception as error:
        raise ImportFileError("The Excel workbook could not be parsed.") from error


def _json_value(value: Any) -> object:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
